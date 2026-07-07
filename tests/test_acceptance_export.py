"""Acceptance suite — the AE1-AE7 examples from the Phase 2 export requirements.

Each test maps to one acceptance example (see
``docs/brainstorms/2026-07-06-xdf-export-module-requirements.md``) and
exercises :func:`~simoscal.export.export_tables` end-to-end. AE1-AE7 need no
real files — the mini fixture (with real decodable bytes) is enough to prove
every degenerate shape and selection/grouping rule, so this suite runs fast
and always-on. One additional real-data pass at the end uses the ``real_cal``
conftest fixture and skips cleanly when the bundled files are absent,
consistent with ``test_acceptance.py``'s convention.

    AE1  2D grid       header/row/values match TableView.values
    AE2  1D shape      no spurious second axis
    AE3  scalar shape  no grid structure, just a value
    AE4  xlsx grouping sheet-per-category, correct membership
    AE5  CSV stacking  single file, ordered, labeled blocks
    AE6  selection     category + explicit symbols union without duplicates
    AE7  direct reuse  render_table() alone matches what the writers produced
"""

from __future__ import annotations

import csv
import struct
from pathlib import Path

import numpy as np
import openpyxl
import pytest

from simoscal import BinImage, CalFile, export_tables, parse_xdf, render_table

FIXTURES = Path(__file__).parent / "fixtures"
MINI_XDF = FIXTURES / "mini.xdf"


@pytest.fixture(scope="module")
def mini_cal() -> CalFile:
    model = parse_xdf(str(MINI_XDF))
    size = model.base_offset + 0x6000
    buf = bytearray(size)
    off = model.base_offset + 0x1000
    buf[off : off + 200] = struct.pack("<100h", *range(100))
    buf[model.base_offset + 0x2000] = 200
    foff = model.base_offset + 0x4000
    buf[foff : foff + 4] = struct.pack("<f", 12.5)
    xoff = model.base_offset + 0x5000
    buf[xoff : xoff + 10] = struct.pack("<5H", 1000, 2000, 3000, 4000, 5000)
    zoff = model.base_offset + 0x5010
    buf[zoff : zoff + 10] = struct.pack("<5H", 10, 20, 30, 40, 50)
    img = BinImage(buf, region_start=model.region_start, region_size=len(buf))
    return CalFile(model, img)


def _read_csv_blocks(path: Path) -> list[list[list[str]]]:
    with open(path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    blocks: list[list[list[str]]] = []
    current: list[list[str]] = []
    for row in rows:
        if row == []:
            if current:
                blocks.append(current)
                current = []
        else:
            current.append(row)
    if current:
        blocks.append(current)
    return blocks


# --------------------------------------------------------------------------- #
# AE1 — 2D grid fidelity
# --------------------------------------------------------------------------- #
def test_ae1_2d_grid_matches_table_values(mini_cal: CalFile, tmp_path):
    view = mini_cal.get("SYM_10X10")
    out = tmp_path / "ae1.csv"
    export_tables(mini_cal, out, symbols=["SYM_10X10"])

    (block,) = _read_csv_blocks(out)
    meta, header, *data_rows = block
    assert meta == ["SYM_10X10", "Ten by Ten", "", "", "%"]
    assert header[1:] == [str(float(i)) for i in range(10)]
    row_labels = [r[0] for r in data_rows]
    assert row_labels == [str(float(i)) for i in range(10)]
    recovered = np.array([[float(c) for c in r[1:]] for r in data_rows])
    np.testing.assert_array_equal(recovered, view.values)


# --------------------------------------------------------------------------- #
# AE2 — 1D shape, no spurious second axis
# --------------------------------------------------------------------------- #
def test_ae2_1d_no_spurious_axis(mini_cal: CalFile, tmp_path):
    out = tmp_path / "ae2.csv"
    export_tables(mini_cal, out, symbols=["PROFILE_1D"])

    (block,) = _read_csv_blocks(out)
    meta, header, data = block
    assert meta[0] == "PROFILE_1D"
    assert header == ["1000.0", "2000.0", "3000.0", "4000.0", "5000.0"]
    assert data == ["10.0", "20.0", "30.0", "40.0", "50.0"]


# --------------------------------------------------------------------------- #
# AE3 — scalar shape, no grid structure
# --------------------------------------------------------------------------- #
def test_ae3_scalar_no_grid_structure(mini_cal: CalFile, tmp_path):
    out = tmp_path / "ae3.csv"
    export_tables(mini_cal, out, symbols=["SYM_SCALAR"])

    (block,) = _read_csv_blocks(out)
    meta, value_row = block
    assert meta[0] == "SYM_SCALAR"
    assert value_row == ["200.0"]


# --------------------------------------------------------------------------- #
# AE4 — xlsx sheet-per-category with correct membership
# --------------------------------------------------------------------------- #
def test_ae4_xlsx_sheet_per_category(mini_cal: CalFile, tmp_path):
    out = tmp_path / "ae4.xlsx"
    export_tables(
        mini_cal, out, symbols=["SYM_10X10", "SYM_SCALAR", "PROFILE_1D"]
    )

    wb = openpyxl.load_workbook(out)
    assert set(wb.sheetnames) == {"Boost Control", "Fuel Trim"}

    def symbols_on(sheet_name: str) -> set[str]:
        ws = wb[sheet_name]
        return {
            row[0]
            for row in ws.iter_rows(values_only=True, min_col=1, max_col=1)
            if row[0] is not None
        } & {"SYM_10X10", "SYM_SCALAR", "PROFILE_1D"}

    assert symbols_on("Boost Control") == {"SYM_10X10", "SYM_SCALAR"}
    assert symbols_on("Fuel Trim") == {"SYM_SCALAR", "PROFILE_1D"}


# --------------------------------------------------------------------------- #
# AE5 — CSV single-file, ordered, labeled blocks
# --------------------------------------------------------------------------- #
def test_ae5_csv_single_file_ordered_blocks(mini_cal: CalFile, tmp_path):
    out = tmp_path / "ae5.csv"
    export_tables(
        mini_cal, out, symbols=["PROFILE_1D", "SYM_10X10", "SYM_SCALAR"]
    )

    blocks = _read_csv_blocks(out)
    assert len(blocks) == 3
    assert [b[0][0] for b in blocks] == ["PROFILE_1D", "SYM_10X10", "SYM_SCALAR"]


# --------------------------------------------------------------------------- #
# AE6 — category + explicit-symbol selection unions without duplicates
# --------------------------------------------------------------------------- #
def test_ae6_selection_union_no_duplicates(mini_cal: CalFile, tmp_path):
    out = tmp_path / "ae6.csv"
    export_tables(mini_cal, out, symbols=["SYM_10X10"], category="Boost Control")

    blocks = _read_csv_blocks(out)
    assert len(blocks) == 2
    assert {b[0][0] for b in blocks} == {"SYM_10X10", "SYM_SCALAR"}


# --------------------------------------------------------------------------- #
# AE7 — direct render_table() reuse matches the writer's numbers
# --------------------------------------------------------------------------- #
def test_ae7_direct_render_matches_writer_output(mini_cal: CalFile, tmp_path):
    view = mini_cal.get("SYM_10X10")
    rt = render_table(view)

    out = tmp_path / "ae7.csv"
    export_tables(mini_cal, out, symbols=["SYM_10X10"])
    (block,) = _read_csv_blocks(out)
    _, header, *data_rows = block

    assert header[1:] == [str(v) for v in rt.x_labels]
    recovered = np.array([[float(c) for c in r[1:]] for r in data_rows])
    np.testing.assert_array_equal(recovered, rt.values)


# --------------------------------------------------------------------------- #
# Error — unrecognized output suffix
# --------------------------------------------------------------------------- #
def test_export_tables_unrecognized_suffix_raises(mini_cal: CalFile, tmp_path):
    out = tmp_path / "ae.txt"
    with pytest.raises(ValueError):
        export_tables(mini_cal, out, symbols=["SYM_10X10"])


# --------------------------------------------------------------------------- #
# Real-data pass — skips cleanly when the bundled files are absent
# --------------------------------------------------------------------------- #
def test_export_real_selection_csv_and_xlsx(real_cal, tmp_path):
    symbols = [v.symbol for v in real_cal.unique_tables()[:5] if v.symbol]

    csv_out = tmp_path / "real.csv"
    export_tables(real_cal, csv_out, symbols=symbols)
    assert len(_read_csv_blocks(csv_out)) == len(symbols)

    xlsx_out = tmp_path / "real.xlsx"
    export_tables(real_cal, xlsx_out, symbols=symbols)
    wb = openpyxl.load_workbook(xlsx_out)
    assert len(wb.sheetnames) > 0
