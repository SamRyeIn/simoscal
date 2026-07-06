"""Tests for the U2-U5 export module (simoscal.export)."""

from __future__ import annotations

import csv
import struct
from pathlib import Path

import numpy as np
import openpyxl
import pytest

from simoscal import (
    AmbiguousTableError,
    BinImage,
    CalFile,
    RenderedTable,
    parse_xdf,
    render_table,
    select_tables,
    write_csv,
    write_xlsx,
)

FIXTURES = Path(__file__).parent / "fixtures"
MINI_XDF = FIXTURES / "mini.xdf"


@pytest.fixture(scope="module")
def mini_cal() -> CalFile:
    model = parse_xdf(str(MINI_XDF))
    # Selection doesn't decode values, so a zero-filled buffer is sufficient —
    # sized just enough to cover the fixture's highest declared address.
    size = model.base_offset + 0x6000
    buf = bytearray(size)
    img = BinImage(buf, region_start=model.region_start, region_size=len(buf))
    return CalFile(model, img)


@pytest.fixture(scope="module")
def mini_cal_with_data() -> CalFile:
    """Same mini.xdf, but with real decodable bytes (mirrors test_render.py)."""
    model = parse_xdf(str(MINI_XDF))
    size = model.base_offset + 0x6000
    buf = bytearray(size)
    off = model.base_offset + 0x1000
    buf[off : off + 200] = struct.pack("<100h", *range(100))
    buf[model.base_offset + 0x2000] = 200
    foff = model.base_offset + 0x4000
    buf[foff : foff + 4] = struct.pack("<f", 12.5)
    xoff = model.base_offset + 0x5000
    buf[xoff : xoff + 10] = struct.pack("<5H", 1000, 2000, 3000, 4000, 5000)
    zoff = model.base_offset + 0x5010
    buf[zoff : zoff + 10] = struct.pack("<5H", 10, 20, 30, 40, 50)
    img = BinImage(buf, region_start=model.region_start, region_size=len(buf))
    return CalFile(model, img)


# --------------------------------------------------------------------------- #
# Happy path
# --------------------------------------------------------------------------- #
def test_select_by_explicit_symbols(mini_cal: CalFile):
    views = select_tables(mini_cal, symbols=["SYM_10X10", "SYM_SCALAR"])
    assert {v.uniqueid for v in views} == {0x100, 0x200}


def test_select_by_category(mini_cal: CalFile):
    views = select_tables(mini_cal, category="Boost Control")
    assert {v.uniqueid for v in views} == {0x100, 0x200}


def test_select_all_tables(mini_cal: CalFile):
    views = select_tables(mini_cal, all_tables=True)
    assert {v.uniqueid for v in views} == {v.uniqueid for v in mini_cal.unique_tables()}
    assert len(views) == 5


# --------------------------------------------------------------------------- #
# Edge — overlap dedup
# --------------------------------------------------------------------------- #
def test_select_symbol_and_category_overlap_dedups(mini_cal: CalFile):
    views = select_tables(mini_cal, symbols=["SYM_10X10"], category="Boost Control")
    assert {v.uniqueid for v in views} == {0x100, 0x200}
    assert len(views) == 2


# --------------------------------------------------------------------------- #
# Errors
# --------------------------------------------------------------------------- #
def test_select_unknown_symbol_raises_keyerror(mini_cal: CalFile):
    with pytest.raises(KeyError):
        select_tables(mini_cal, symbols=["DOES_NOT_EXIST"])


def test_select_ambiguous_symbol_raises(mini_cal: CalFile):
    with pytest.raises(AmbiguousTableError):
        select_tables(mini_cal, symbols=["SYM_DUP"])


def test_select_no_input_raises_valueerror(mini_cal: CalFile):
    with pytest.raises(ValueError):
        select_tables(mini_cal)


# --------------------------------------------------------------------------- #
# Integration — real data category selection
# --------------------------------------------------------------------------- #
def test_select_real_axis_category_count(real_cal):
    views = select_tables(real_cal, category="Axis")
    assert len(views) == 444


# --------------------------------------------------------------------------- #
# U3 — CSV writer
# --------------------------------------------------------------------------- #
def _read_csv_blocks(path: Path) -> list[list[list[str]]]:
    """Split a written CSV into blank-line-separated blocks of rows."""
    with open(path, newline="") as f:
        rows = list(csv.reader(f))
    blocks: list[list[list[str]]] = []
    current: list[list[str]] = []
    for row in rows:
        if row == []:
            if current:
                blocks.append(current)
                current = []
        else:
            current.append(row)
    if current:
        blocks.append(current)
    return blocks


def test_write_csv_10x10_roundtrips(mini_cal_with_data: CalFile, tmp_path):
    view = mini_cal_with_data.get("SYM_10X10")
    rt = render_table(view)
    out = tmp_path / "out.csv"
    write_csv([rt], out)

    blocks = _read_csv_blocks(out)
    assert len(blocks) == 1
    meta, *grid = blocks[0]
    assert meta == ["SYM_10X10", "Ten by Ten", "%"]
    header, *data_rows = grid
    assert header[0] == ""
    recovered = np.array([[float(c) for c in r[1:]] for r in data_rows])
    np.testing.assert_array_equal(recovered, view.values)


def test_write_csv_multiple_tables_stacked_in_order(mini_cal_with_data: CalFile, tmp_path):
    tables = [
        render_table(mini_cal_with_data.get("SYM_10X10")),
        render_table(mini_cal_with_data.get("PROFILE_1D")),
        render_table(mini_cal_with_data.get("SYM_SCALAR")),
    ]
    out = tmp_path / "out.csv"
    write_csv(tables, out)

    blocks = _read_csv_blocks(out)
    assert len(blocks) == 3
    assert [b[0][0] for b in blocks] == ["SYM_10X10", "PROFILE_1D", "SYM_SCALAR"]

    # 1D block: no leading blank cell in header, no leading label in data row.
    _, header, data = blocks[1]
    assert header == ["1000.0", "2000.0", "3000.0", "4000.0", "5000.0"]
    assert data == ["10.0", "20.0", "30.0", "40.0", "50.0"]

    # Scalar block: bare value line only, no header row.
    _, value_row = blocks[2]
    assert value_row == ["200.0"]


def test_write_csv_full_precision_value_roundtrips(mini_cal_with_data: CalFile, tmp_path):
    view = mini_cal_with_data.get("SYM_10X10")
    rt = render_table(view)
    out = tmp_path / "out.csv"
    write_csv([rt], out)

    blocks = _read_csv_blocks(out)
    _, header, *data_rows = blocks[0]
    recovered = np.array([[float(c) for c in r[1:]] for r in data_rows])
    np.testing.assert_array_equal(recovered, view.values)
    # A cell with a many-decimal-digit value round-trips exactly through str->float.
    assert float(data_rows[1][2]) == view.values[1, 1]


def test_write_csv_real_tables_roundtrip(real_cal, tmp_path):
    tables = [render_table(v) for v in real_cal.unique_tables()[:5]]
    out = tmp_path / "out.csv"
    write_csv(tables, out)

    blocks = _read_csv_blocks(out)
    assert len(blocks) == 5
    for rt, block in zip(tables, blocks):
        assert block[0][0] == (rt.symbol or "")


# --------------------------------------------------------------------------- #
# U4 — xlsx writer
# --------------------------------------------------------------------------- #
def _strip_trailing_none(row: tuple) -> tuple:
    """Drop right-padding ``None``s from a row (sheet-wide max_column padding)."""
    row = list(row)
    while row and row[-1] is None:
        row.pop()
    return tuple(row)


def _sheet_blocks(ws) -> list[list[tuple]]:
    """Split a worksheet into blank-row-separated blocks (mirrors _read_csv_blocks)."""
    blocks: list[list[tuple]] = []
    current: list[tuple] = []
    for row in ws.iter_rows(values_only=True):
        row = _strip_trailing_none(row)
        if not row:
            if current:
                blocks.append(current)
                current = []
        else:
            current.append(row)
    if current:
        blocks.append(current)
    return blocks


def test_write_xlsx_sheets_by_category(mini_cal_with_data: CalFile, tmp_path):
    tables = [
        render_table(mini_cal_with_data.get("SYM_10X10")),
        render_table(mini_cal_with_data.get("SYM_SCALAR")),
        render_table(mini_cal_with_data.get("PROFILE_1D")),
    ]
    out = tmp_path / "out.xlsx"
    write_xlsx(tables, out)

    wb = openpyxl.load_workbook(out)
    assert set(wb.sheetnames) == {"Boost Control", "Fuel Trim"}

    boost_symbols = {b[0][0] for b in _sheet_blocks(wb["Boost Control"])}
    assert boost_symbols == {"SYM_10X10", "SYM_SCALAR"}

    fuel_symbols = {b[0][0] for b in _sheet_blocks(wb["Fuel Trim"])}
    assert fuel_symbols == {"SYM_SCALAR", "PROFILE_1D"}


def test_write_xlsx_multi_category_table_identical_on_both_sheets(
    mini_cal_with_data: CalFile, tmp_path
):
    tables = [
        render_table(mini_cal_with_data.get("SYM_10X10")),
        render_table(mini_cal_with_data.get("SYM_SCALAR")),
        render_table(mini_cal_with_data.get("PROFILE_1D")),
    ]
    out = tmp_path / "out.xlsx"
    write_xlsx(tables, out)

    wb = openpyxl.load_workbook(out)
    boost_block = next(
        b for b in _sheet_blocks(wb["Boost Control"]) if b[0][0] == "SYM_SCALAR"
    )
    fuel_block = next(
        b for b in _sheet_blocks(wb["Fuel Trim"]) if b[0][0] == "SYM_SCALAR"
    )
    assert boost_block == fuel_block


def test_write_xlsx_sheet_name_sanitized_and_deduped(tmp_path):
    long_prefix = "X" * 40
    rt1 = RenderedTable(
        symbol="A", title="A title", units="u",
        categories=(long_prefix + "_one",),
        x_labels=(0.0,), y_labels=None, values=np.array([[1.0]]),
    )
    rt2 = RenderedTable(
        symbol="B", title="B title", units="u",
        categories=(long_prefix + "_two",),
        x_labels=(0.0,), y_labels=None, values=np.array([[2.0]]),
    )
    out = tmp_path / "out.xlsx"
    write_xlsx([rt1, rt2], out)

    wb = openpyxl.load_workbook(out)
    assert len(wb.sheetnames) == 2
    assert wb.sheetnames[0] != wb.sheetnames[1]
    assert all(len(name) <= 31 for name in wb.sheetnames)


def test_write_xlsx_real_axis_category_slice(real_cal, tmp_path):
    views = select_tables(real_cal, category="Axis")
    tables = [render_table(v) for v in views]
    out = tmp_path / "out.xlsx"
    write_xlsx(tables, out)

    wb = openpyxl.load_workbook(out)
    assert "Axis" in wb.sheetnames
    assert len(_sheet_blocks(wb["Axis"])) == 444
